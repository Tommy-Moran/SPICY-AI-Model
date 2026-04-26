"""SPICY-AI vs Physician vs GPT-4o — full analysis (n=50, 150 plans).

Reproducibility design
----------------------
This script supports two input modes:

1. **CSV mode (default for external reviewers)** — reads the de-identified
   long-format scores from `data/scores_long_v3.csv`. This is the path NEJM
   AI reviewers, Zenodo users, and anyone cloning the public GitHub repo
   should use. No raw clinical data required.

2. **Source-xlsx mode (authors only)** — if `--source-xlsx` is passed and
   the two scoring workbooks are present (paths configurable via flags or
   `SPICY_XLSX_DIR` env var), the script unblinds them via the per-workbook
   allocation keys and rebuilds the long-format CSV from scratch.

Either mode produces an identical `data/scores_long_v3.csv` and identical
downstream statistics and figures.

Statistical methods
-------------------
- Pairwise Mann-Whitney U (two-sided) for ordinal Likert scores.
- Holm-Bonferroni adjusted p-values across the 12 score comparisons
  (4 domains x 3 pairs). Both raw and adjusted p-values are reported.
- Fisher exact for minor/major error rates (3 pairs each).
- Means, SDs, medians, IQRs reported for descriptive transparency; inference
  is driven by the rank-based tests.

Pre-specified allocation: A = Physician, B = SPICY-AI, C = GPT-4o.

Run:
    python3 analysis_v3.py                  # CSV mode (default)
    python3 analysis_v3.py --source-xlsx    # rebuild CSV from raw xlsx
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

import pandas as pd
import numpy as np
from scipy import stats as sps
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch, FancyArrowPatch
from matplotlib.lines import Line2D

# ── Paths (all relative to this script) ──────────────────────────────────────
HERE       = Path(__file__).resolve().parent
DATA_DIR   = HERE / "data"
FIG_DIR    = HERE / "figures"
SCORES_CSV = DATA_DIR / "scores_long_v3.csv"

DATA_DIR.mkdir(exist_ok=True)
FIG_DIR.mkdir(exist_ok=True)

GROUPS  = ["Physician", "SPICY-AI", "GPT-4o"]
DOMAINS = [("diag", "Diagnostic accuracy"),
           ("mgmt", "Management appropriateness"),
           ("comp", "Completeness"),
           ("clar", "Clarity")]
COLORS  = {"Physician": "#4C72B0", "SPICY-AI": "#55A868", "GPT-4o": "#C44E52"}
PAIRS   = [("Physician", "SPICY-AI"),
           ("Physician", "GPT-4o"),
           ("SPICY-AI", "GPT-4o")]


# ── Source-xlsx mode (authors only) ──────────────────────────────────────────
def load_scored_wb(path: Path, alloc_sheet_name: str) -> list[dict]:
    """Unblind a scoring workbook via its allocation key sheet."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws, ak = wb["Scoring"], wb[alloc_sheet_name]

    alloc: dict[int, dict[str, str]] = {}
    for row in ak.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            alloc[int(row[0])] = {"A": row[1], "B": row[2], "C": row[3]}
        except (TypeError, ValueError):
            continue

    records, current_rin = [], None
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0]
        if val is None:
            continue
        if isinstance(val, (int, float)) or (isinstance(val, str) and str(val).strip().isdigit()):
            current_rin = int(float(val))
            continue
        if str(val).strip() in ("A", "B", "C") and current_rin in alloc:
            plan = str(val).strip()
            diag, mgmt, comp, clar, minor, major = row[2:8]
            if any(not isinstance(s, (int, float)) for s in (diag, mgmt, comp, clar, minor, major)):
                continue
            records.append(dict(
                rin=current_rin, label=plan, group=alloc[current_rin][plan],
                diag=int(diag), mgmt=int(mgmt), comp=int(comp), clar=int(clar),
                minor=int(minor), major=int(major),
            ))
    return records


def rebuild_from_xlsx(xlsx_dir: Path) -> pd.DataFrame:
    wb1 = xlsx_dir / "CSANZ_Cases_Scored_1-10.xlsx"
    wb2 = xlsx_dir / "CSANZ_Cases_Scored_11-50.xlsx"
    if not wb1.exists() or not wb2.exists():
        sys.exit(f"--source-xlsx mode: workbooks not found in {xlsx_dir}\n"
                 f"  expected: {wb1.name} and {wb2.name}")
    recs = (load_scored_wb(wb1, "Allocation Key (Cases 1-10)") +
            load_scored_wb(wb2, "Allocation Key (DO NOT OPEN WHILE SCORING)"))
    df = pd.DataFrame(recs)
    df.to_csv(SCORES_CSV, index=False)
    print(f"Rebuilt {SCORES_CSV} from raw xlsx ({len(df)} rows)")
    return df


# ── Holm-Bonferroni helper ──────────────────────────────────────────────────
def holm_bonferroni(pvals: list[float]) -> list[float]:
    """Step-down Holm adjustment. Returns adjusted p-values in input order."""
    m = len(pvals)
    order = sorted(range(m), key=lambda i: pvals[i])
    adj = [0.0] * m
    running = 0.0
    for rank, idx in enumerate(order):
        running = max(running, pvals[idx] * (m - rank))
        adj[idx] = min(running, 1.0)
    return adj


# ── Main ────────────────────────────────────────────────────────────────────
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--source-xlsx", action="store_true",
                    help="Rebuild scores_long_v3.csv from raw scoring workbooks "
                         "(authors only; requires the two .xlsx files).")
    ap.add_argument("--xlsx-dir", type=Path,
                    default=Path(os.environ.get("SPICY_XLSX_DIR", HERE.parent)),
                    help="Directory containing CSANZ_Cases_Scored_*.xlsx "
                         "(default: parent of this script, or $SPICY_XLSX_DIR).")
    args = ap.parse_args()

    if args.source_xlsx:
        df = rebuild_from_xlsx(args.xlsx_dir)
    else:
        if not SCORES_CSV.exists():
            sys.exit(f"Missing {SCORES_CSV}. Run with --source-xlsx to rebuild "
                     f"from raw workbooks, or fetch the de-identified CSV from the repo.")
        df = pd.read_csv(SCORES_CSV)

    # ── Validation ──────────────────────────────────────────────────────────
    n_plans = len(df)
    group_counts = df.groupby("group").size().to_dict()
    print(f"Loaded {n_plans} plan rows; group counts: {group_counts}")
    assert n_plans == 150, f"Expected 150 rows, got {n_plans}"
    assert group_counts == {"Physician": 50, "SPICY-AI": 50, "GPT-4o": 50}, \
        f"Unexpected group counts: {group_counts}"

    # ── Summary statistics ──────────────────────────────────────────────────
    summary = []
    for dom, _ in DOMAINS:
        for g in GROUPS:
            v = df.loc[df.group == g, dom]
            summary.append(dict(domain=dom, group=g, n=len(v),
                                mean=v.mean(), sd=v.std(ddof=1),
                                median=v.median(),
                                q1=v.quantile(0.25), q3=v.quantile(0.75)))
    sdf = pd.DataFrame(summary)
    sdf.to_csv(DATA_DIR / "summary_v3.csv", index=False)
    print("\n=== Summary statistics (n=50/group) ===")
    print(sdf.to_string(index=False))

    # ── Mann-Whitney U with Holm-Bonferroni adjustment ──────────────────────
    mw_rows = []
    for dom, label in DOMAINS:
        for a, b in PAIRS:
            x = df.loc[df.group == a, dom].values
            y = df.loc[df.group == b, dom].values
            u, p = sps.mannwhitneyu(x, y, alternative="two-sided")
            mw_rows.append(dict(domain=dom, label=label, groupA=a, groupB=b,
                                medA=float(np.median(x)), medB=float(np.median(y)),
                                meanA=float(np.mean(x)), meanB=float(np.mean(y)),
                                U=u, p=p))
    mw = pd.DataFrame(mw_rows)
    mw["p_holm"] = holm_bonferroni(mw["p"].tolist())
    mw.to_csv(DATA_DIR / "mannwhitney_v3.csv", index=False)

    print("\n=== Mann-Whitney U (pairwise, two-sided; Holm-Bonferroni across 12 tests) ===")
    print(mw[["domain", "groupA", "groupB", "meanA", "meanB", "U", "p", "p_holm"]].to_string(index=False))

    # ── Error rates (Fisher exact) ──────────────────────────────────────────
    def error_tests(col: str, n_per_group: int = 50) -> list[dict]:
        rows = []
        for a, b in PAIRS:
            xa = int(df.loc[df.group == a, col].sum())
            xb = int(df.loc[df.group == b, col].sum())
            _, p = sps.fisher_exact([[xa, n_per_group - xa], [xb, n_per_group - xb]])
            rows.append(dict(error=col, groupA=a, rateA=xa / n_per_group,
                             groupB=b, rateB=xb / n_per_group, p_fisher=p))
        return rows

    err = pd.DataFrame(error_tests("minor") + error_tests("major"))
    err.to_csv(DATA_DIR / "error_rates_v3.csv", index=False)
    print("\n=== Error rates (Fisher exact, pairwise) ===")
    print(err.to_string(index=False))

    err_summary = df.groupby("group")[["minor", "major"]].agg(["sum", "mean"])
    err_summary.to_csv(DATA_DIR / "error_summary_v3.csv")
    print("\n=== Error rates by group ===")
    print(err_summary)

    # ── Figures ─────────────────────────────────────────────────────────────
    make_figure_1a(df, mw)
    make_figure_1b(df)
    make_figure_2(df)
    make_figure_3()

    # ── Stats summary text ──────────────────────────────────────────────────
    with open(DATA_DIR / "stats_output_v3.txt", "w") as f:
        f.write("=== SPICY-AI extended statistics (cases 1-50, n=50/group) ===\n\n")
        f.write("Total plans: 150 (50 per group)\n\n")
        f.write("Summary statistics:\n")
        f.write(sdf.to_string(index=False) + "\n\n")
        f.write("Mann-Whitney U pairwise (two-sided), Holm-Bonferroni adjusted across 12 tests:\n")
        f.write(mw[["domain", "groupA", "groupB", "meanA", "meanB", "U", "p", "p_holm"]].to_string(index=False) + "\n\n")
        f.write("Error rates (Fisher exact, pairwise):\n")
        f.write(err.to_string(index=False) + "\n\n")
        f.write("Error rates by group:\n")
        f.write(err_summary.to_string() + "\n\n")
        f.write("Note: study is not powered to detect differences in major-error rates "
                "(observed counts 1-2 per arm); the null Fisher results should be interpreted "
                "as a feasibility/safety signal, not as evidence of equivalence.\n")

    print(f"\nAll outputs written under {DATA_DIR} and {FIG_DIR}")


# ── Figure 1A ───────────────────────────────────────────────────────────────
def make_figure_1a(df: pd.DataFrame, mw: pd.DataFrame) -> None:
    fig, ax = plt.subplots(figsize=(11, 7.5))
    x, width = np.arange(len(DOMAINS)), 0.25
    for i, g in enumerate(GROUPS):
        means = [df.loc[df.group == g, d[0]].mean() for d in DOMAINS]
        sds = [df.loc[df.group == g, d[0]].std(ddof=1) for d in DOMAINS]
        ax.bar(x + (i - 1) * width, means, width, yerr=sds, capsize=4,
               label=g, color=COLORS[g], edgecolor="black", linewidth=0.7)

    ax.set_xticks(x)
    ax.set_xticklabels([d[1] for d in DOMAINS], fontsize=11)
    ax.set_ylabel("Mean Likert score (1-5)", fontsize=12)
    ax.set_ylim(0, 7.6)
    ax.set_yticks([0, 1, 2, 3, 4, 5])
    ax.legend(loc="upper left", bbox_to_anchor=(0.0, 0.98),
              frameon=False, ncol=3, fontsize=10)
    ax.set_title("Figure 1A. Mean Likert scores by domain (n=50 per group). "
                 "Brackets show Holm-adjusted p-values: *** p<0.001, ** p<0.01, * p<0.05",
                 fontsize=11, loc="left", pad=12)

    def bracket(x1, x2, y, p, h=0.08):
        if p >= 0.05:
            return
        label = "***" if p < 0.001 else "**" if p < 0.01 else "*"
        ax.plot([x1, x1, x2, x2], [y, y + h, y + h, y], lw=1.0, c="black")
        ax.text((x1 + x2) / 2, y + h + 0.02, label, ha="center", fontsize=10)

    offs = {"Physician": -width, "SPICY-AI": 0.0, "GPT-4o": width}
    order = {("Physician", "SPICY-AI"): 0, ("SPICY-AI", "GPT-4o"): 1, ("Physician", "GPT-4o"): 2}
    for di, (dom, _) in enumerate(DOMAINS):
        sub = mw[mw.domain == dom].reset_index(drop=True)
        top = max(df.loc[df.group == g, dom].mean() + df.loc[df.group == g, dom].std(ddof=1) for g in GROUPS)
        y0 = max(top + 0.25, 5.35)
        for _, r in sub.iterrows():
            idx = order[(r.groupA, r.groupB)]
            bracket(di + offs[r.groupA], di + offs[r.groupB], y0 + idx * 0.42, r.p_holm)

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    for ext in ("png", "pdf", "tiff"):
        fig.savefig(FIG_DIR / f"Figure1A.{ext}", dpi=300, bbox_inches="tight")
    plt.close(fig)
    print("Saved Figure 1A")


# ── Figure 1B ───────────────────────────────────────────────────────────────
def make_figure_1b(df: pd.DataFrame) -> None:
    fig, ax = plt.subplots(figsize=(9, 4.8))
    err_rates = (df.groupby("group")[["minor", "major"]].mean() * 100).reindex(GROUPS)
    y, h = np.arange(len(GROUPS)), 0.38
    bars_minor = ax.barh(y - h / 2, err_rates["minor"], h, label="Minor errors",
                         color="#E1A95F", edgecolor="black", linewidth=0.7)
    bars_major = ax.barh(y + h / 2, err_rates["major"], h, label="Major errors",
                         color="#B03030", edgecolor="black", linewidth=0.7)
    for bars in (bars_minor, bars_major):
        for b in bars:
            w = b.get_width()
            ax.text(w + 0.5, b.get_y() + b.get_height() / 2, f"{w:.1f}%", va="center", fontsize=10)
    ax.set_yticks(y)
    ax.set_yticklabels(GROUPS)
    ax.set_xlabel("Error rate (% of plans)", fontsize=12)
    ax.set_xlim(0, max(err_rates.values.max() * 1.3, 25))
    ax.legend(loc="lower right", frameon=False)
    ax.set_title("Figure 1B. Minor and major error rates by group", fontsize=12, loc="left")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    for ext in ("png", "pdf", "tiff"):
        fig.savefig(FIG_DIR / f"Figure1B.{ext}", dpi=300, bbox_inches="tight")
    plt.close(fig)
    print("Saved Figure 1B")


# ── Figure 2 ────────────────────────────────────────────────────────────────
def make_figure_2(df: pd.DataFrame) -> None:
    fig, axes = plt.subplots(1, 4, figsize=(14, 5), sharey=True)
    rng = np.random.default_rng(seed=42)
    for ax, (dom, label) in zip(axes, DOMAINS):
        data = [df.loc[df.group == g, dom].values for g in GROUPS]
        bp = ax.boxplot(data, tick_labels=GROUPS, patch_artist=True, widths=0.55,
                        medianprops=dict(color="black", linewidth=1.5),
                        flierprops=dict(marker="o", markersize=4, alpha=0.6))
        for patch, g in zip(bp["boxes"], GROUPS):
            patch.set_facecolor(COLORS[g])
            patch.set_alpha(0.85)
            patch.set_edgecolor("black")
        for i, d in enumerate(data):
            xs = rng.normal(i + 1, 0.07, len(d))
            ax.scatter(xs, d, s=10, alpha=0.35, color="black", zorder=3)
        ax.set_title(label, fontsize=11)
        ax.set_ylim(0.5, 5.5)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.tick_params(axis="x", labelrotation=15)
    axes[0].set_ylabel("Likert score (1-5)", fontsize=12)
    fig.suptitle("Figure 2. Likert score distributions by domain (n=50 per group)",
                 fontsize=13, y=1.00, x=0.07, ha="left")
    fig.tight_layout()
    for ext in ("png", "pdf", "tiff"):
        fig.savefig(FIG_DIR / f"Figure2.{ext}", dpi=300, bbox_inches="tight")
    plt.close(fig)
    print("Saved Figure 2")


# ── Figure 3 (pipeline diagram) ─────────────────────────────────────────────
def make_figure_3() -> None:
    fig, ax = plt.subplots(figsize=(14, 7.5))
    ax.set_xlim(0, 14); ax.set_ylim(0, 8); ax.axis("off")

    def box(x, y, w, h, text, color="#E3ECF7", edge="#2F5F8F", fontsize=10, bold=False):
        b = FancyBboxPatch((x, y), w, h, boxstyle="round,pad=0.08,rounding_size=0.18",
                           fc=color, ec=edge, linewidth=1.6)
        ax.add_patch(b)
        fw = "bold" if bold else "normal"
        ax.text(x + w / 2, y + h / 2, text, ha="center", va="center",
                fontsize=fontsize, fontweight=fw, wrap=True)

    def arrow(x1, y1, x2, y2, color="#2F5F8F", style="-|>", rad=0.0, lw=1.6):
        cs = f"arc3,rad={rad}" if rad else "arc3"
        ax.add_patch(FancyArrowPatch((x1, y1), (x2, y2), arrowstyle=style,
                                      mutation_scale=14, connectionstyle=cs,
                                      color=color, linewidth=lw))

    TOP_Y, TOP_H = 5.4, 1.3
    MID_Y, MID_H = 3.0, 1.3
    EVAL_Y, EVAL_H = 0.5, 1.2

    ax.text(0.05, TOP_Y + TOP_H / 2, "Input →\nModel", ha="left", va="center",
            fontsize=9, color="#666666", style="italic")
    ax.text(0.05, MID_Y + MID_H / 2, "Reasoning →\nOutput", ha="left", va="center",
            fontsize=9, color="#666666", style="italic")
    ax.text(0.05, EVAL_Y + EVAL_H / 2, "Evaluation", ha="left", va="center",
            fontsize=9, color="#666666", style="italic")

    box(1.2, TOP_Y, 2.6, TOP_H, "Clinical context\n(RACPC referral:\nHx, Ix, exam,\nrisk factors)",
        color="#EEF3F8", edge="#2F5F8F", fontsize=9)
    box(4.2, TOP_Y, 2.4, TOP_H, "De-identification\n& prompt\nassembly",
        color="#FFF3D6", edge="#B5832A", fontsize=9)
    box(7.0, TOP_Y, 2.8, TOP_H, "DeepSeek-R1 14B\n(locally deployed,\nOllama, on-prem)",
        color="#D9ECD9", edge="#3F7B3F", fontsize=9, bold=True)
    box(10.2, TOP_Y, 3.2, TOP_H, "Retrieval-Augmented\nGeneration\n• CSANZ/ESC/AHA guidelines\n• Local formulary & pathways",
        color="#FAD9D9", edge="#B03030", fontsize=9)

    box(1.2, MID_Y, 3.2, MID_H, "Structured chain-of-\nthought reasoning\n(Dx → Ix → Mx → F/U)",
        color="#E7E0F7", edge="#5C3E9F", fontsize=9)
    box(4.8, MID_Y, 3.6, MID_H, "Management plan output\n(risk optimisation, Ix,\nreferrals, follow-up)",
        color="#EEF3F8", edge="#2F5F8F", fontsize=9, bold=True)
    box(8.8, MID_Y, 2.6, MID_H, "Clinician review\n& validation",
        color="#FFE8B3", edge="#B5832A", fontsize=9)
    box(11.8, MID_Y, 1.6, MID_H, "Final\nmanagement\nplan",
        color="#EEF3F8", edge="#2F5F8F", fontsize=9)

    box(1.2, EVAL_Y, 12.2, EVAL_H,
        "Blinded Likert evaluation   (50 RACPC cases × 3 plan sources = 150 plans)\n"
        "Diagnostic accuracy · Management appropriateness · Completeness · Clarity  (1–5)"
        "   +   minor / major error flags",
        color="#F5F5F5", edge="#555555", fontsize=9)

    arrow(3.8, TOP_Y + TOP_H / 2, 4.2, TOP_Y + TOP_H / 2)
    arrow(6.6, TOP_Y + TOP_H / 2, 7.0, TOP_Y + TOP_H / 2)
    arrow(10.2, TOP_Y + TOP_H / 2, 9.8, TOP_Y + TOP_H / 2, color="#B03030")
    arrow(8.4, TOP_Y, 2.8, MID_Y + MID_H, rad=-0.2)
    arrow(4.4, MID_Y + MID_H / 2, 4.8, MID_Y + MID_H / 2)
    arrow(8.4, MID_Y + MID_H / 2, 8.8, MID_Y + MID_H / 2)
    arrow(11.4, MID_Y + MID_H / 2, 11.8, MID_Y + MID_H / 2)
    arrow(6.6, MID_Y, 6.6, EVAL_Y + EVAL_H)
    arrow(12.6, MID_Y, 12.6, EVAL_Y + EVAL_H)
    arrow(2.5, TOP_Y, 2.5, EVAL_Y + EVAL_H, rad=-0.25, color="#666666", lw=1.2)

    legend_elems = [
        Line2D([0], [0], color="#2F5F8F", lw=2, label="Data / reasoning flow"),
        Line2D([0], [0], color="#B03030", lw=2, label="Retrieved knowledge"),
        Line2D([0], [0], color="#666666", lw=2, label="Reference input to evaluation"),
    ]
    ax.legend(handles=legend_elems, loc="lower right",
              bbox_to_anchor=(0.99, 0.00), frameon=False, fontsize=8.5)
    ax.text(7.0, 7.55, "Figure 3. SPICY-AI clinical reasoning pipeline",
            ha="center", fontsize=14, fontweight="bold")

    fig.tight_layout()
    for ext in ("png", "pdf", "tiff"):
        fig.savefig(FIG_DIR / f"Figure3_pipeline.{ext}", dpi=300, bbox_inches="tight")
    plt.close(fig)
    print("Saved Figure 3 (pipeline)")


if __name__ == "__main__":
    main()
